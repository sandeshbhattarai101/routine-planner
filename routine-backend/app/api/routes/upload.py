import csv
import io

from fastapi import APIRouter
from fastapi import Depends
from fastapi import UploadFile
from fastapi import File
from fastapi import HTTPException
from openpyxl import load_workbook

from app.core.permissions import school_admin_only
from app.core.dependencies import get_current_school_id

router = APIRouter(
    prefix="/uploads",
    tags=["Uploads"]
)

ALLOWED_EXTENSIONS = ["xlsx", "csv"]


def parse_sheet_rows(sheet) -> list[dict]:
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header = [str(col) if col is not None else "" for col in next(rows_iter)]
    except StopIteration:
        return []

    rows = []
    for row in rows_iter:
        if all(value is None for value in row):
            continue
        rows.append({header[i]: row[i] for i in range(len(header))})

    return rows


def parse_xlsx(contents: bytes) -> list[dict]:
    """Parses every sheet in the workbook into {name, columns, rows}."""
    workbook = load_workbook(filename=io.BytesIO(contents), read_only=True, data_only=True)

    sheets = []
    for sheet_name in workbook.sheetnames:
        rows = parse_sheet_rows(workbook[sheet_name])
        if not rows:
            continue
        sheets.append({
            "name": sheet_name,
            "columns": list(rows[0].keys()),
            "rows": rows,
        })

    return sheets


def parse_csv(contents: bytes, filename: str) -> list[dict]:
    text = contents.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    rows = [dict(row) for row in reader]

    if not rows:
        return []

    return [{
        "name": filename.rsplit(".", 1)[0],
        "columns": list(rows[0].keys()),
        "rows": rows,
    }]


@router.post("/")
async def upload_excel(
    file: UploadFile = File(...),
    school_id=Depends(get_current_school_id),
    _=Depends(school_admin_only)
):
    extension = file.filename.split(".")[-1].lower()

    if extension not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail="Only .xlsx and .csv files are allowed"
        )

    contents = await file.read()

    try:
        if extension == "xlsx":
            sheets = parse_xlsx(contents)
        else:
            sheets = parse_csv(contents, file.filename)
    except Exception:
        raise HTTPException(
            status_code=400,
            detail="Could not parse file. Please check the file format."
        )

    if not sheets:
        raise HTTPException(
            status_code=400,
            detail="File contains no data rows."
        )

    return {
        "filename": file.filename,
        "sheets": sheets,
    }